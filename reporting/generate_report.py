import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

DB_PATH = Path(__file__).parent.parent / "netflix.db"
OUTPUT_PATH = Path(__file__).parent.parent / "Netflix_Report.xlsx"

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F2937")
LABEL_FONT = Font(name="Arial", bold=True, size=11)
BASE_FONT = Font(name="Arial", size=11)

RAW_COLUMNS = [
    "show_id", "type", "title", "primary_country", "date_added",
    "release_year", "rating", "duration_value", "duration_unit",
    "primary_genre",
]


def fetch_data() -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql(f"SELECT {', '.join(RAW_COLUMNS)} FROM titles", conn)
    conn.close()
    df["date_added"] = pd.to_datetime(df["date_added"], errors="coerce")
    return df


def style_header_row(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def write_raw_data_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title = "Raw Data"

    ws.append(RAW_COLUMNS)
    style_header_row(ws, 1, len(RAW_COLUMNS))

    for _, row in df.iterrows():
        ws.append(list(row))

    date_col = RAW_COLUMNS.index("date_added") + 1
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=date_col).number_format = "yyyy-mm-dd"

    for col_idx, col_name in enumerate(RAW_COLUMNS, start=1):
        width = max(12, min(30, len(col_name) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    return ws


def write_summary_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Summary")

    ws["A1"] = "Netflix Content Summary Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    n_raw_rows = len(df) + 1  

    # --- Table 1: Content type breakdown ------------------------------
    ws["A3"] = "Content Type Breakdown"
    ws["A3"].font = LABEL_FONT

    ws.append([])  
    ws["A5"], ws["B5"], ws["C5"] = "Type", "Count", "Share"
    style_header_row(ws, 5, 3)

    ws["A6"] = "Movie"
    ws["B6"] = f'=COUNTIF(\'Raw Data\'!B2:B{n_raw_rows},"Movie")'
    ws["A7"] = "TV Show"
    ws["B7"] = f'=COUNTIF(\'Raw Data\'!B2:B{n_raw_rows},"TV Show")'
    ws["A8"] = "Total"
    ws["B8"] = "=SUM(B6:B7)"
    ws["A8"].font = LABEL_FONT
    ws["B8"].font = LABEL_FONT

    for r in (6, 7, 8):
        ws[f"C{r}"] = f"=B{r}/$B$8"
        ws[f"C{r}"].number_format = "0.0%"

    # --- Table 2: Top 10 genres ----------------------------------------
    genre_counts = (
        df["primary_genre"].value_counts().head(10).reset_index()
    )
    genre_counts.columns = ["genre", "count"]

    start_row = 11
    ws.cell(row=start_row - 1, column=1, value="Top 10 Genres").font = LABEL_FONT
    ws.cell(row=start_row, column=1, value="Genre")
    ws.cell(row=start_row, column=2, value="Count")
    style_header_row(ws, start_row, 2)

    for i, genre in enumerate(genre_counts["genre"], start=1):
        r = start_row + i
        ws.cell(row=r, column=1, value=genre)
        safe_genre = genre.replace('"', '""')
        ws.cell(
            row=r, column=2,
            value=f'=COUNTIF(\'Raw Data\'!J2:J{n_raw_rows},"{safe_genre}")',
        )

    genre_table_end = start_row + len(genre_counts)

    # --- Table 3: Titles added per year (2008-2021) --------------------
    years = list(range(2008, 2022))
    year_start_row = genre_table_end + 3

    ws.cell(row=year_start_row - 1, column=1, value="Titles Added by Year").font = LABEL_FONT
    ws.cell(row=year_start_row, column=1, value="Year")
    ws.cell(row=year_start_row, column=2, value="Titles Added")
    style_header_row(ws, year_start_row, 2)

    date_col_letter = "E" 
    for i, year in enumerate(years, start=1):
        r = year_start_row + i
        ws.cell(row=r, column=1, value=year)
        ws.cell(
            row=r, column=2,
            value=(
                f"=COUNTIFS('Raw Data'!{date_col_letter}2:{date_col_letter}{n_raw_rows},"
                f'">="&DATE({year},1,1),'
                f"'Raw Data'!{date_col_letter}2:{date_col_letter}{n_raw_rows},"
                f'"<"&DATE({year + 1},1,1))'
            ),
        )

    year_table_end = year_start_row + len(years)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10

    return ws, {
        "type_table": (5, 8),
        "genre_table": (start_row, genre_table_end),
        "year_table": (year_start_row, year_table_end),
    }


def add_charts(ws, table_ranges: dict):
    # Pie chart: content type breakdown
    pie = PieChart()
    pie.title = "Movies vs TV Shows"
    labels = Reference(ws, min_col=1, min_row=6, max_row=7)
    data = Reference(ws, min_col=2, min_row=5, max_row=7)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width, pie.height = 9, 7
    ws.add_chart(pie, "E5")

    # Bar chart: top genres
    g_start, g_end = table_ranges["genre_table"]
    bar = BarChart()
    bar.type = "bar"
    bar.title = "Top 10 Genres"
    bar.y_axis.title = "Genre"
    bar.x_axis.title = "Number of Titles"
    labels = Reference(ws, min_col=1, min_row=g_start + 1, max_row=g_end)
    data = Reference(ws, min_col=2, min_row=g_start, max_row=g_end)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.width, bar.height = 15, 10
    ws.add_chart(bar, "E22")

    # Line chart: titles added per year
    y_start, y_end = table_ranges["year_table"]
    line = LineChart()
    line.title = "Titles Added to Netflix by Year"
    line.y_axis.title = "Titles Added"
    line.x_axis.title = "Year"
    labels = Reference(ws, min_col=1, min_row=y_start + 1, max_row=y_end)
    data = Reference(ws, min_col=2, min_row=y_start, max_row=y_end)
    line.add_data(data, titles_from_data=True)
    line.set_categories(labels)
    line.width, line.height = 15, 10
    ws.add_chart(line, "E45")


def main():
    print("[report] Fetching data from netflix.db ...")
    df = fetch_data()

    wb = Workbook()
    write_raw_data_sheet(wb, df)
    summary_ws, table_ranges = write_summary_sheet(wb, df)
    add_charts(summary_ws, table_ranges)

    wb.save(OUTPUT_PATH)
    print(f"[report] Saved report to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
